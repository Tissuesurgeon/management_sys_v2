from __future__ import annotations

import calendar
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.http import Http404, HttpRequest, HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from workforce.decorators import org_admin_required, worker_required
from workforce.forms import (
    MaintenanceTaskForm,
    PasswordResetWithCodeForm,
    ProfilePhotoForm,
    SignupForm,
    TaskStateUpdateForm,
    UserAccountForm,
    WorkerInvitationForm,
    WorkerProfileForm,
    maintenance_task_bulk_formset_factory,
)
from workforce.models import (
    MaintenanceTask,
    Profile,
    TaskEvidencePhoto,
    TaskState,
    Worker,
    WorkerInvitation,
    WorkerPasswordResetCode,
    generate_unique_password_reset_code,
)
from workforce.utils import ensure_profile
from workforce.services.date_utils import date_key, parse_date_key
try:
    from PIL import Image as PILImage
except ImportError:  # pragma: no cover
    PILImage = None

_EVIDENCE_MAX_BYTES = 5 * 1024 * 1024
_EVIDENCE_MAX_PER_TASK = 20


def _validate_evidence_upload(f) -> None:
    if not f or not getattr(f, 'name', ''):
        raise ValidationError('Choose an image file.')
    if getattr(f, 'size', 0) > _EVIDENCE_MAX_BYTES:
        raise ValidationError('Image must be under 5 MB.')
    if PILImage is not None:
        f.seek(0)
        try:
            with PILImage.open(f) as im:
                im.verify()
        except Exception as exc:  # noqa: BLE001
            raise ValidationError('Please upload a valid image file (JPEG, PNG, or WebP).') from exc
        f.seek(0)


def _sync_evidence_photo_count(derived_id: str) -> int:
    n = TaskEvidencePhoto.objects.filter(derived_task_id=derived_id).count()
    st = TaskState.objects.filter(derived_task_id=derived_id).first()
    if st:
        if st.photo_count != n:
            st.photo_count = n
            st.save(update_fields=['photo_count', 'last_saved_at'])
    elif n:
        TaskState.objects.create(
            derived_task_id=derived_id,
            status=TaskState.Status.SCHEDULED,
            notes='',
            checklist_done={},
            photo_count=n,
        )
    return n


from workforce.services.tasks import (
    checklist_is_complete,
    collect_state_map_for_ids,
    effective_worker_display_status,
    generate_tasks_from_calendar,
    merge_task_state,
    parse_derived_task_id,
    trade_label,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore[misc, assignment]
    Font = None  # type: ignore[misc, assignment]
    get_column_letter = None  # type: ignore[misc, assignment]

def _profile(request):
    return ensure_profile(request.user)


def _active_merged_tasks(
    events: List[MaintenanceTask],
    day: date,
    *,
    worker_trade: Optional[str] = None,
    viewer_worker_pk: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """Generated tasks for ``day`` with TaskState merged in; drops only truly completed tasks."""
    tasks = generate_tasks_from_calendar(
        events,
        day,
        worker_trade=worker_trade,
        viewer_worker_pk=viewer_worker_pk,
    )
    ids = [t.id for t in tasks]
    state_map = collect_state_map_for_ids(ids)
    merged = merge_task_state(tasks, state_map)
    return [r for r in merged if r.get('status') != TaskState.Status.COMPLETED]


def _enrich_task_states_for_report(states: List[TaskState]) -> List[Dict[str, Any]]:
    """Join persisted task rows to calendar events and workers for admin reporting."""
    eids: List[int] = []
    for st in states:
        eid, _ = parse_derived_task_id(st.derived_task_id)
        if eid is not None:
            eids.append(eid)
    events = {e.pk: e for e in MaintenanceTask.objects.filter(pk__in=set(eids))}
    rows: List[Dict[str, Any]] = []
    for st in states:
        eid, task_date = parse_derived_task_id(st.derived_task_id)
        ev = events.get(eid) if eid is not None else None
        checklist = list(ev.checklist or []) if ev else []
        done = st.checklist_done or {}
        n_done = sum(
            1
            for i in range(len(checklist))
            if done.get(str(i)) or done.get(i)
        )
        rows.append(
            {
                'state': st,
                'event': ev,
                'task_date': task_date,
                'worker_name': trade_label(ev.assigned_trade) if ev else '—',
                'checklist_done_count': n_done,
                'checklist_total': len(checklist),
                'title': ev.title if ev else 'Unknown or deleted event',
                'location': (ev.location or '') if ev else '',
            },
        )
    return rows


def _completed_task_states_for_worker(worker: Worker):
    """Persisted task rows for this worker with status completed (via calendar event assignment)."""
    eids = list(
        MaintenanceTask.objects.filter(assigned_trade=worker.trade).values_list('pk', flat=True),
    )
    base = TaskState.objects.filter(status=TaskState.Status.COMPLETED).order_by('-last_saved_at')
    if not eids:
        return base.none()
    q = Q()
    for eid in eids:
        q |= Q(derived_task_id__startswith=f'{eid}::')
    return base.filter(q)


def _local_task_start_date_key(ev: MaintenanceTask) -> str:
    s = ev.start
    if timezone.is_naive(s):
        s = timezone.make_aware(s, timezone.get_current_timezone())
    return date_key(timezone.localtime(s).date())


def _upcoming_events_for_worker(worker: Worker, limit: int = 5) -> List[Dict[str, Any]]:
    """Calendar rows for profile: not yet ended, ordered by start; status queued vs in progress."""
    now = timezone.now()
    rows: List[Dict[str, Any]] = []
    for ev in MaintenanceTask.objects.filter(assigned_trade=worker.trade).order_by('start')[:120]:
        start = ev.start
        if timezone.is_naive(start):
            start = timezone.make_aware(start, timezone.get_current_timezone())
        end = start + timedelta(minutes=ev.duration_minutes)
        if end <= now:
            continue
        if now < start:
            status_key = 'queued'
            status_label = 'QUEUED'
        else:
            status_key = 'in_progress'
            status_label = 'IN PROGRESS'
        rows.append(
            {
                'event': ev,
                'status_key': status_key,
                'status_label': status_label,
                'date_key': _local_task_start_date_key(ev),
            },
        )
        if len(rows) >= limit:
            break
    return rows


def _admin_profile_stats() -> Dict[str, int]:
    today = timezone.localdate()
    events_list = list(MaintenanceTask.objects.all())
    today_tasks_count = len(_active_merged_tasks(events_list, today))
    return {
        'worker_count': Worker.objects.count(),
        'today_tasks_count': today_tasks_count,
        'completed_today': TaskState.objects.filter(
            status=TaskState.Status.COMPLETED,
            last_saved_at__date=today,
        ).count(),
    }


def home(request: HttpRequest) -> HttpResponse:
    if not request.user.is_authenticated:
        return redirect('workforce:login')
    prof = _profile(request)
    if prof.role == Profile.Role.ORG_ADMIN:
        return redirect('workforce:admin_dashboard')
    return redirect('workforce:worker_my_tasks')


def login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect('workforce:home')
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        login(request, form.get_user())
        return redirect('workforce:home')
    return render(request, 'workforce/login.html', {'form': form})


def signup(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect('workforce:home')
    form = SignupForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        role = form.cleaned_data['role']
        invitation = getattr(form, '_invitation', None)
        with transaction.atomic():
            inv_locked = None
            if role == Profile.Role.WORKER and invitation:
                inv_locked = WorkerInvitation.objects.select_for_update().get(pk=invitation.pk)
                if inv_locked.claimed_at:
                    messages.error(
                        request,
                        'This registration code was already used. Ask your facility manager for a new one.',
                    )
                    return render(request, 'workforce/signup.html', {'form': form})
            user = form.save(commit=False)
            if form.cleaned_data.get('email'):
                user.email = form.cleaned_data['email']
            user.save()
            Profile.objects.create(user=user, role=role)
            if role == Profile.Role.WORKER and inv_locked:
                Worker.objects.create(
                    user=user,
                    name=inv_locked.name,
                    title=inv_locked.title,
                    gender=inv_locked.gender or '',
                    department=inv_locked.department,
                    employee_id=inv_locked.employee_id or (form.cleaned_data.get('employee_id') or '').strip(),
                    facility_location=inv_locked.facility_location,
                    trade=form.cleaned_data['trade'],
                )
                inv_locked.claimed_at = timezone.now()
                inv_locked.claimed_by = user
                inv_locked.save()
            elif role == Profile.Role.WORKER:
                Worker.objects.create(
                    user=user,
                    name=form.cleaned_data['worker_name'].strip(),
                    employee_id=form.cleaned_data.get('employee_id') or '',
                    trade=form.cleaned_data['trade'],
                )
        login(request, user)
        messages.success(request, 'Account created.')
        return redirect('workforce:home')
    return render(request, 'workforce/signup.html', {'form': form})


def logout_view(request: HttpRequest) -> HttpResponse:
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('workforce:login')


def password_reset_with_code(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect('workforce:home')
    form = PasswordResetWithCodeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form._user
        row = form._reset_row
        with transaction.atomic():
            locked = WorkerPasswordResetCode.objects.select_for_update().get(pk=row.pk)
            if locked.used_at or locked.expires_at <= timezone.now():
                messages.error(
                    request,
                    'This code is no longer valid. Ask your facility manager for a new one.',
                )
                return render(request, 'workforce/password_reset.html', {'form': form})
            user.set_password(form.cleaned_data['password1'])
            user.save()
            locked.used_at = timezone.now()
            locked.save()
        messages.success(request, 'Your password has been reset. You can sign in now.')
        return redirect('workforce:login')
    return render(request, 'workforce/password_reset.html', {'form': form})


@org_admin_required
def admin_dashboard(request: HttpRequest) -> HttpResponse:
    today = timezone.localdate()
    worker_count = Worker.objects.count()
    events_list = list(MaintenanceTask.objects.all())
    today_tasks_count = len(_active_merged_tasks(events_list, today))
    completed_today = TaskState.objects.filter(
        status=TaskState.Status.COMPLETED,
        last_saved_at__date=today,
    ).count()
    open_items = TaskState.objects.exclude(status=TaskState.Status.COMPLETED).count()
    recent_states = list(TaskState.objects.order_by('-last_saved_at')[:6])
    recent_activity = _enrich_task_states_for_report(recent_states)
    return render(
        request,
        'workforce/admin/dashboard.html',
        {
            'worker_count': worker_count,
            'today_tasks_count': today_tasks_count,
            'completed_today': completed_today,
            'open_items': open_items,
            'today': today,
            'recent_activity': recent_activity,
        },
    )


@org_admin_required
def admin_calendar_list(request: HttpRequest) -> HttpResponse:
    """Calendar page — layout parity with Management_sys AdminCalendar + CalendarMonth."""
    now = timezone.localtime()
    today = now.date()

    try:
        y = int(request.GET.get('year') or now.year)
        m = int(request.GET.get('month') or now.month)
    except ValueError:
        y, m = now.year, now.month
    if m < 1:
        m, y = 12, y - 1
    if m > 12:
        m, y = 1, y + 1

    selected_raw = request.GET.get('selected')
    if selected_raw:
        try:
            parts = [int(x) for x in selected_raw.split('-')[:3]]
            selected_date = date(parts[0], parts[1], parts[2])
        except (ValueError, IndexError):
            selected_date = today
    else:
        selected_date = today

    events_qs = MaintenanceTask.objects.all()
    events_list = list(events_qs)
    event_count = len(events_list)

    first = date(y, m, 1)
    # JS Date.getDay(): Sun=0 … Sat=6. Python weekday(): Mon=0 … Sun=6
    start_pad = (first.weekday() + 1) % 7
    days_in_month = calendar.monthrange(y, m)[1]
    cells: List[Optional[date]] = [None] * start_pad
    for d in range(1, days_in_month + 1):
        cells.append(date(y, m, d))
    while len(cells) % 7 != 0:
        cells.append(None)

    tasks_by_day: Dict[str, int] = {}
    for d in range(1, days_in_month + 1):
        day = date(y, m, d)
        tasks_by_day[date_key(day)] = len(_active_merged_tasks(events_list, day))

    calendar_cells: List[Dict[str, Any]] = []
    for cell in cells:
        if cell is None:
            calendar_cells.append({'pad': True})
        else:
            dk = date_key(cell)
            calendar_cells.append(
                {
                    'pad': False,
                    'date': cell,
                    'key': dk,
                    'count': tasks_by_day.get(dk, 0),
                },
            )

    selected_key = date_key(selected_date)
    today_key = date_key(today)

    day_events: List[MaintenanceTask] = []
    for ev in events_list:
        if _local_task_start_date_key(ev) == selected_key:
            day_events.append(ev)

    tasks_for_day = [r['task'] for r in _active_merged_tasks(events_list, selected_date)]

    prev_m, prev_y = (m - 1, y) if m > 1 else (12, y - 1)
    next_m, next_y = (m + 1, y) if m < 12 else (1, y + 1)

    header_subtitle = (
        f'{selected_date.strftime("%B %Y")} · Maintenance tasks on the calendar'
    )
    return render(
        request,
        'workforce/admin/calendar_list.html',
        {
            'events': events_list,
            'event_count': event_count,
            'view_year': y,
            'view_month': m,
            'month_title': date(y, m, 1).strftime('%B %Y'),
            'calendar_cells': calendar_cells,
            'selected_date': selected_date,
            'selected_key': selected_key,
            'today_key': today_key,
            'day_events': day_events,
            'tasks_for_day': tasks_for_day,
            'header_subtitle': header_subtitle,
            'tasks_for_day_count': len(tasks_for_day),
            'prev_year': prev_y,
            'prev_month': prev_m,
            'next_year': next_y,
            'next_month': next_m,
        },
    )


@org_admin_required
def admin_calendar_create(request: HttpRequest) -> HttpResponse:
    form = MaintenanceTaskForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Maintenance task saved.')
        return redirect('workforce:admin_calendar_list')
    return render(request, 'workforce/admin/calendar_form.html', {'form': form, 'title': 'New maintenance task'})


@org_admin_required
def admin_calendar_edit(request: HttpRequest, pk: int) -> HttpResponse:
    ev = get_object_or_404(MaintenanceTask, pk=pk)
    form = MaintenanceTaskForm(request.POST or None, instance=ev)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Maintenance task updated.')
        return redirect('workforce:admin_calendar_list')
    return render(
        request,
        'workforce/admin/calendar_form.html',
        {'form': form, 'title': 'Edit maintenance task', 'event': ev},
    )


@org_admin_required
def admin_calendar_delete(request: HttpRequest, pk: int) -> HttpResponse:
    ev = get_object_or_404(MaintenanceTask, pk=pk)
    if request.method == 'POST':
        ev.delete()
        messages.success(request, 'Maintenance task removed.')
        return redirect('workforce:admin_calendar_list')
    return render(request, 'workforce/admin/calendar_confirm_delete.html', {'event': ev})


@org_admin_required
def admin_calendar_month_bulk(request: HttpRequest) -> HttpResponse:
    """Create multiple maintenance tasks for a single calendar month in one submit."""
    now = timezone.localtime()
    today = now.date()

    if request.method == 'POST':
        try:
            y = int(request.POST.get('bulk_year') or now.year)
            m = int(request.POST.get('bulk_month') or now.month)
        except (TypeError, ValueError):
            y, m = now.year, now.month
    else:
        try:
            y = int(request.GET.get('year') or now.year)
            m = int(request.GET.get('month') or now.month)
        except (TypeError, ValueError):
            y, m = now.year, now.month

    if m < 1:
        m, y = 12, y - 1
    if m > 12:
        m, y = 1, y + 1

    first_day = date(y, m, 1)
    last_day = date(y, m, calendar.monthrange(y, m)[1])

    FormSet = maintenance_task_bulk_formset_factory(first_day, last_day, extra=1)
    prev_m, prev_y = (m - 1, y) if m > 1 else (12, y - 1)
    next_m, next_y = (m + 1, y) if m < 12 else (1, y + 1)

    if request.method == 'POST':
        formset = FormSet(request.POST)
        if formset.is_valid():
            rows_to_save = [
                f
                for f in formset
                if (f.cleaned_data.get('title') or '').strip()
            ]
            if not rows_to_save:
                messages.error(request, 'Add at least one task with a title.')
            else:
                with transaction.atomic():
                    for f in rows_to_save:
                        cd = f.cleaned_data
                        rtype = cd['recurrence_type']
                        rec_end = (
                            cd.get('recurrence_end')
                            if rtype != MaintenanceTask.Recurrence.NONE
                            else None
                        )
                        MaintenanceTask.objects.create(
                            title=cd['title'].strip(),
                            description=(cd.get('description') or '').strip(),
                            location=(cd.get('location') or '').strip(),
                            start=cd['start'],
                            duration_minutes=cd['duration_minutes'],
                            assigned_trade=cd['assigned_trade'],
                            recurrence_type=rtype,
                            recurrence_end=rec_end,
                            checklist=cd['checklist_text'],
                            color=cd['color'],
                        )
                messages.success(
                    request,
                    f'Created {len(rows_to_save)} maintenance task(s) for {first_day.strftime("%B %Y")}.',
                )
                qs = urlencode(
                    {
                        'year': y,
                        'month': m,
                        'selected': first_day.isoformat(),
                    },
                )
                return redirect(f"{reverse('workforce:admin_calendar_list')}?{qs}")
        # invalid formset or empty rows message: fall through
    else:
        formset = FormSet()

    month_label = first_day.strftime('%B %Y')
    return render(
        request,
        'workforce/admin/calendar_month_bulk.html',
        {
            'formset': formset,
            'view_year': y,
            'view_month': m,
            'month_label': month_label,
            'first_day': first_day,
            'last_day': last_day,
            'prev_year': prev_y,
            'prev_month': prev_m,
            'next_year': next_y,
            'next_month': next_m,
            'today': today,
        },
    )


@org_admin_required
def admin_tasks(request: HttpRequest) -> HttpResponse:
    raw = request.GET.get('date') or ''
    try:
        day = datetime.strptime(raw, '%Y-%m-%d').date() if raw else timezone.localdate()
    except ValueError:
        day = timezone.localdate()
    worker_pk = request.GET.get('worker')
    w_pk: Optional[int] = None
    if worker_pk not in (None, ''):
        try:
            w_pk = int(worker_pk)
        except ValueError:
            w_pk = None

    events = list(MaintenanceTask.objects.all())
    w_trade: Optional[str] = None
    v_pk: Optional[int] = None
    if w_pk is not None:
        filt_w = get_object_or_404(Worker, pk=w_pk)
        w_trade = filt_w.trade
        v_pk = w_pk
    merged = _active_merged_tasks(events, day, worker_trade=w_trade, viewer_worker_pk=v_pk)
    for row in merged:
        row['worker_name'] = trade_label(row['task'].assigned_trade)
    workers = Worker.objects.all()
    return render(
        request,
        'workforce/admin/tasks.html',
        {
            'day': day,
            'date_str': day.isoformat(),
            'merged_tasks': merged,
            'workers': workers,
            'filter_worker': w_pk,
        },
    )


@org_admin_required
def admin_task_detail(request: HttpRequest, event_id: int, date_key: str) -> HttpResponse:
    ev = get_object_or_404(MaintenanceTask, pk=event_id)
    try:
        day = parse_date_key(date_key)
    except (ValueError, IndexError):
        raise Http404('Invalid date') from None
    tasks = generate_tasks_from_calendar([ev], day, worker_trade=None, viewer_worker_pk=None)
    if not tasks:
        raise Http404('No task on this date for this event.')
    derived_id = tasks[0].id
    state_map = collect_state_map_for_ids([derived_id])
    merged_rows = merge_task_state(tasks, state_map)
    row = merged_rows[0]
    return render(
        request,
        'workforce/admin/task_detail.html',
        {
            'event': ev,
            'task': tasks[0],
            'row': row,
            'date_key': date_key,
            'persist': row.get('persist'),
            'worker_name': trade_label(ev.assigned_trade),
            'edit_url': reverse('workforce:admin_calendar_edit', args=[ev.pk]),
        },
    )


@org_admin_required
def admin_task_report(request: HttpRequest) -> HttpResponse:
    """Detailed report of persisted task work (status, notes, photos, checklist)."""
    today = timezone.localdate()
    raw_from = request.GET.get('date_from') or ''
    raw_to = request.GET.get('date_to') or ''
    try:
        date_from = datetime.strptime(raw_from, '%Y-%m-%d').date() if raw_from else today - timedelta(days=30)
    except ValueError:
        date_from = today - timedelta(days=30)
    try:
        date_to = datetime.strptime(raw_to, '%Y-%m-%d').date() if raw_to else today
    except ValueError:
        date_to = today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    status_filter = request.GET.get('status') or ''
    valid_statuses = {c for c, _ in TaskState.Status.choices}
    if status_filter not in valid_statuses:
        status_filter = ''

    worker_pk = request.GET.get('worker')
    w_pk: Optional[int] = None
    if worker_pk not in (None, ''):
        try:
            w_pk = int(worker_pk)
        except ValueError:
            w_pk = None

    qs = TaskState.objects.filter(
        last_saved_at__date__gte=date_from,
        last_saved_at__date__lte=date_to,
    ).order_by('-last_saved_at')

    if status_filter:
        qs = qs.filter(status=status_filter)

    if w_pk is not None:
        fw = get_object_or_404(Worker, pk=w_pk)
        eids = list(
            MaintenanceTask.objects.filter(assigned_trade=fw.trade).values_list('pk', flat=True),
        )
        if not eids:
            qs = qs.none()
        else:
            q = Q()
            for eid in eids:
                q |= Q(derived_task_id__startswith=f'{eid}::')
            qs = qs.filter(q)

    summary = {
        'total': qs.count(),
        'completed': qs.filter(status=TaskState.Status.COMPLETED).count(),
        'in_progress': qs.filter(status=TaskState.Status.IN_PROGRESS).count(),
        'scheduled': qs.filter(status=TaskState.Status.SCHEDULED).count(),
    }

    paginator = Paginator(qs, 30)
    page = request.GET.get('page') or 1
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    rows = _enrich_task_states_for_report(list(page_obj.object_list))
    workers = Worker.objects.all()
    q = request.GET.copy()
    q.pop('page', None)
    pagination_qs = q.urlencode()
    return render(
        request,
        'workforce/admin/task_report.html',
        {
            'rows': rows,
            'page_obj': page_obj,
            'paginator': paginator,
            'summary': summary,
            'date_from': date_from,
            'date_to': date_to,
            'date_from_str': date_from.isoformat(),
            'date_to_str': date_to.isoformat(),
            'status_filter': status_filter,
            'workers': workers,
            'filter_worker': w_pk,
            'pagination_qs': pagination_qs,
        },
    )


@org_admin_required
def admin_workers(request: HttpRequest) -> HttpResponse:
    trade_counts = {
        row['assigned_trade']: row['n']
        for row in MaintenanceTask.objects.values('assigned_trade').annotate(n=Count('id'))
    }
    workers = list(Worker.objects.order_by('name'))
    for w in workers:
        w.event_count = trade_counts.get(w.trade, 0)
    return render(request, 'workforce/admin/workers.html', {'workers': workers})


@org_admin_required
def admin_worker_invites(request: HttpRequest) -> HttpResponse:
    invites = WorkerInvitation.objects.select_related('created_by', 'claimed_by').order_by('-created_at')
    return render(request, 'workforce/admin/worker_invites.html', {'invites': invites})


@org_admin_required
def admin_worker_invite_create(request: HttpRequest) -> HttpResponse:
    form = WorkerInvitationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        inv = form.save(commit=False)
        inv.created_by = request.user
        inv.save()
        messages.success(
            request,
            f'Registration created. Share code {inv.invite_code} with {inv.name}.',
        )
        return redirect('workforce:admin_worker_invites')
    return render(request, 'workforce/admin/worker_invite_form.html', {'form': form})


@org_admin_required
def admin_worker_reset_code_generate(request: HttpRequest, pk: int) -> HttpResponse:
    worker = get_object_or_404(Worker, pk=pk)
    if request.method != 'POST':
        return redirect('workforce:admin_worker_detail', pk=pk)
    if not worker.user_id:
        messages.error(request, 'This worker has no login account yet.')
        return redirect('workforce:admin_worker_detail', pk=pk)
    prof = Profile.objects.filter(user_id=worker.user_id).first()
    if not prof or prof.role != Profile.Role.WORKER:
        messages.error(
            request,
            'Password reset codes can only be issued for worker accounts.',
        )
        return redirect('workforce:admin_worker_detail', pk=pk)
    now = timezone.now()
    WorkerPasswordResetCode.objects.filter(
        user_id=worker.user_id,
        used_at__isnull=True,
        expires_at__gt=now,
    ).delete()
    code = generate_unique_password_reset_code()
    expires = now + timedelta(hours=48)
    WorkerPasswordResetCode.objects.create(
        user=worker.user,
        code=code,
        expires_at=expires,
        created_by=request.user,
    )
    messages.success(
        request,
        f'Password reset code for {worker.user.username}: {code}. '
        f'Expires {timezone.localtime(expires).strftime("%b %d, %Y %I:%M %p")}. '
        'Share this code with the worker securely; it works once.',
    )
    return redirect('workforce:admin_worker_detail', pk=pk)


@org_admin_required
def admin_worker_detail(request: HttpRequest, pk: int) -> HttpResponse:
    worker = get_object_or_404(Worker, pk=pk)
    today = timezone.localdate()
    months_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]

    completed_qs = _completed_task_states_for_worker(worker)
    completed_count = completed_qs.count()
    paginator = Paginator(completed_qs, 25)
    page = request.GET.get('page') or 1
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    completed_rows = _enrich_task_states_for_report(list(page_obj.object_list))

    return render(
        request,
        'workforce/admin/worker_detail.html',
        {
            'worker': worker,
            'export_default_year': today.year,
            'export_default_month': today.month,
            'export_month_choices': months_choices,
            'completed_rows': completed_rows,
            'completed_count': completed_count,
            'page_obj': page_obj,
            'paginator': paginator,
        },
    )


@org_admin_required
def admin_worker_completed_export(request: HttpRequest, pk: int) -> HttpResponse:
    """Excel export of completed tasks for one worker in a calendar month (by completion date)."""
    if Workbook is None:
        return HttpResponseBadRequest('openpyxl is not installed.')
    worker = get_object_or_404(Worker, pk=pk)
    try:
        year = int(request.GET.get('year') or timezone.localdate().year)
        month = int(request.GET.get('month') or timezone.localdate().month)
    except (TypeError, ValueError):
        return HttpResponseBadRequest('Invalid year or month.')
    if month < 1 or month > 12 or year < 2000 or year > 2100:
        return HttpResponseBadRequest('Year or month out of range.')

    last_d = calendar.monthrange(year, month)[1]
    start_d = date(year, month, 1)
    end_d = date(year, month, last_d)

    qs = (
        _completed_task_states_for_worker(worker)
        .filter(
            last_saved_at__date__gte=start_d,
            last_saved_at__date__lte=end_d,
        )
        .order_by('last_saved_at')
    )
    states = list(qs)
    rows_data = _enrich_task_states_for_report(states)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Completed tasks'

    headers = [
        'Derived task ID',
        'Task title',
        'Trade',
        'Occurrence date',
        'Location',
        'Completed at',
        'Notes',
        'Photos',
        'Checklist done',
        'Checklist total',
    ]
    bold = Font(bold=True) if Font else None
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        if bold:
            cell.font = bold

    for r, item in enumerate(rows_data, start=2):
        st: TaskState = item['state']
        completed_at = (
            timezone.localtime(st.last_saved_at).strftime('%Y-%m-%d %H:%M')
            if st.last_saved_at
            else ''
        )
        occ = item.get('task_date') or ''
        if occ and len(occ) == 10:
            try:
                occ_fmt = datetime.strptime(occ, '%Y-%m-%d').date().isoformat()
            except ValueError:
                occ_fmt = occ
        else:
            occ_fmt = occ
        ws.cell(row=r, column=1, value=st.derived_task_id)
        ws.cell(row=r, column=2, value=item.get('title') or '')
        ws.cell(row=r, column=3, value=item.get('worker_name') or '')
        ws.cell(row=r, column=4, value=occ_fmt)
        ws.cell(row=r, column=5, value=item.get('location') or '')
        ws.cell(row=r, column=6, value=completed_at)
        ws.cell(row=r, column=7, value=(st.notes or '').strip() or '—')
        ws.cell(row=r, column=8, value=st.photo_count)
        ws.cell(row=r, column=9, value=item.get('checklist_done_count', 0))
        ws.cell(row=r, column=10, value=item.get('checklist_total', 0))

    col_widths = [30, 36, 22, 14, 28, 18, 48, 10, 14, 14]
    if get_column_letter:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    slug = re.sub(r'[^a-zA-Z0-9._-]+', '_', worker.name.strip())[:48] or f'worker_{worker.pk}'
    fname = f'completed_tasks_{slug}_{year}-{month:02d}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@worker_required
def worker_my_tasks(request: HttpRequest) -> HttpResponse:
    worker = get_object_or_404(Worker, user=request.user)
    day_raw = request.GET.get('date') or ''
    try:
        day = datetime.strptime(day_raw, '%Y-%m-%d').date() if day_raw else timezone.localdate()
    except ValueError:
        day = timezone.localdate()

    events = list(MaintenanceTask.objects.all())
    merged = _active_merged_tasks(
        events,
        day,
        worker_trade=worker.trade,
        viewer_worker_pk=worker.pk,
    )
    return render(
        request,
        'workforce/worker/my_tasks.html',
        {'merged_tasks': merged, 'day': day, 'date_str': day.isoformat()},
    )


@worker_required
def worker_schedule(request: HttpRequest) -> HttpResponse:
    worker = get_object_or_404(Worker, user=request.user)
    now = timezone.localdate()
    try:
        y = int(request.GET.get('year') or now.year)
        m = int(request.GET.get('month') or now.month)
    except ValueError:
        y, m = now.year, now.month
    if m < 1:
        m, y = 12, y - 1
    if m > 12:
        m, y = 1, y + 1

    cal = calendar.Calendar(firstweekday=0)
    weeks = cal.monthdatescalendar(y, m)
    events = list(MaintenanceTask.objects.all())

    grid: List[List[Dict[str, Any]]] = []
    for week in weeks:
        row = []
        for d in week:
            if d.month != m:
                row.append({'day': None, 'in_month': False, 'count': 0, 'date_str': ''})
                continue
            n_active = len(
                _active_merged_tasks(
                    events,
                    d,
                    worker_trade=worker.trade,
                    viewer_worker_pk=worker.pk,
                ),
            )
            row.append(
                {
                    'day': d.day,
                    'in_month': True,
                    'count': n_active,
                    'date_str': d.isoformat(),
                },
            )
        grid.append(row)

    prev_m, prev_y = (m - 1, y) if m > 1 else (12, y - 1)
    next_m, next_y = (m + 1, y) if m < 12 else (1, y + 1)

    month_task_count = 0
    last_d = calendar.monthrange(y, m)[1]
    for d in range(1, last_d + 1):
        day = date(y, m, d)
        month_task_count += len(
            _active_merged_tasks(
                events,
                day,
                worker_trade=worker.trade,
                viewer_worker_pk=worker.pk,
            ),
        )

    subtitle = f'{calendar.month_name[m]} {y} · {month_task_count} tasks this month (you)'
    return render(
        request,
        'workforce/worker/schedule.html',
        {
            'year': y,
            'month': m,
            'month_name': calendar.month_name[m],
            'weeks': grid,
            'prev_year': prev_y,
            'prev_month': prev_m,
            'next_year': next_y,
            'next_month': next_m,
            'month_task_count': month_task_count,
            'subtitle': subtitle,
        },
    )


@worker_required
def worker_history(request: HttpRequest) -> HttpResponse:
    worker = get_object_or_404(Worker, user=request.user)
    states = TaskState.objects.filter(status=TaskState.Status.COMPLETED).order_by('-last_saved_at')
    rows: List[Dict[str, Any]] = []
    for st in states:
        parts = st.derived_task_id.split('::')
        if len(parts) != 2:
            continue
        try:
            eid = int(parts[0])
        except ValueError:
            continue
        ev = MaintenanceTask.objects.filter(pk=eid, assigned_trade=worker.trade).first()
        if not ev:
            continue
        checklist = list(ev.checklist or [])
        if not checklist_is_complete(checklist, st.checklist_done):
            continue
        rows.append({'state': st, 'event': ev, 'date_key': parts[1]})
    return render(request, 'workforce/worker/history.html', {'rows': rows})


@org_admin_required
def admin_profile(request: HttpRequest) -> HttpResponse:
    user = request.user
    prof = ensure_profile(user)
    user_form = UserAccountForm(request.POST or None, instance=user)
    photo_form = ProfilePhotoForm(
        request.POST or None,
        request.FILES or None,
        instance=prof,
    )
    if request.method == 'POST':
        u_ok = user_form.is_valid()
        p_ok = photo_form.is_valid()
        if u_ok and p_ok:
            user_form.save()
            photo_form.save()
            messages.success(request, 'Profile updated.')
            return redirect('workforce:admin_profile')
    return render(
        request,
        'workforce/admin/profile.html',
        {
            'user_form': user_form,
            'photo_form': photo_form,
            'profile': prof,
            'profile_stats': _admin_profile_stats(),
        },
    )


@worker_required
def worker_profile(request: HttpRequest) -> HttpResponse:
    worker = get_object_or_404(Worker, user=request.user)
    user = request.user
    prof = ensure_profile(user)
    user_form = UserAccountForm(request.POST or None, instance=user)
    photo_form = ProfilePhotoForm(
        request.POST or None,
        request.FILES or None,
        instance=prof,
    )
    worker_form = WorkerProfileForm(request.POST or None, instance=worker)
    if request.method == 'POST':
        u_ok = user_form.is_valid()
        p_ok = photo_form.is_valid()
        w_ok = worker_form.is_valid()
        if u_ok and p_ok and w_ok:
            user_form.save()
            photo_form.save()
            worker_form.save()
            messages.success(request, 'Profile updated.')
            return redirect('workforce:worker_profile')
    return render(
        request,
        'workforce/worker/profile.html',
        {
            'user_form': user_form,
            'photo_form': photo_form,
            'worker_form': worker_form,
            'worker': worker,
            'profile': prof,
            'upcoming_events': _upcoming_events_for_worker(worker),
        },
    )


@worker_required
def worker_task_detail(request: HttpRequest, event_id: int, date_key: str) -> HttpResponse:
    worker = get_object_or_404(Worker, user=request.user)
    ev = get_object_or_404(MaintenanceTask, pk=event_id, assigned_trade=worker.trade)
    try:
        day = parse_date_key(date_key)
    except (ValueError, IndexError):
        raise Http404('Invalid date') from None

    tasks = generate_tasks_from_calendar(
        [ev],
        day,
        worker_trade=worker.trade,
        viewer_worker_pk=worker.pk,
    )
    if not tasks:
        raise Http404('No task on this date for this event.')
    task = tasks[0]
    derived_id = task.id
    st = TaskState.objects.filter(derived_task_id=derived_id).first()

    checklist_keys = list(task.checklist or [])
    initial: Dict[str, Any] = {}
    if st:
        initial['status'] = st.status
        initial['notes'] = st.notes
    else:
        initial['status'] = TaskState.Status.SCHEDULED
        initial['notes'] = ''

    form = TaskStateUpdateForm(
        request.POST or None,
        initial=initial,
        checklist_keys=checklist_keys,
        checklist_done=st.checklist_done if st else {},
    )

    if request.method == 'POST':
        remove_pk = request.POST.get('remove_photo')
        if remove_pk:
            try:
                rid = int(remove_pk)
            except (TypeError, ValueError):
                raise Http404('Invalid photo') from None
            photo = get_object_or_404(TaskEvidencePhoto, pk=rid, derived_task_id=derived_id)
            photo.image.delete(save=False)
            photo.delete()
            _sync_evidence_photo_count(derived_id)
            messages.success(request, 'Photo removed.')
            return redirect('workforce:worker_task_detail', event_id=event_id, date_key=date_key)

        action = request.POST.get('action', 'save')
        if action == 'upload_photo':
            f = request.FILES.get('evidence_photo')
            try:
                _validate_evidence_upload(f)
            except ValidationError as exc:
                msg = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
                messages.error(request, msg)
                return redirect('workforce:worker_task_detail', event_id=event_id, date_key=date_key)
            if TaskEvidencePhoto.objects.filter(derived_task_id=derived_id).count() >= _EVIDENCE_MAX_PER_TASK:
                messages.error(
                    request,
                    f'You can upload at most {_EVIDENCE_MAX_PER_TASK} photos per task.',
                )
                return redirect('workforce:worker_task_detail', event_id=event_id, date_key=date_key)
            TaskEvidencePhoto.objects.create(derived_task_id=derived_id, image=f)
            _sync_evidence_photo_count(derived_id)
            messages.success(request, 'Photo uploaded.')
            return redirect('workforce:worker_task_detail', event_id=event_id, date_key=date_key)

    if request.method == 'POST' and form.is_valid():
        action = request.POST.get('action', 'save')
        done: Dict[str, bool] = {}
        for i, _label in enumerate(checklist_keys):
            field = f'check_{i}'
            if field in form.fields and form.cleaned_data.get(field):
                done[str(i)] = True
        status_val = form.cleaned_data['status']
        if action == 'complete':
            if checklist_is_complete(checklist_keys, done):
                status_val = TaskState.Status.COMPLETED
            else:
                status_val = TaskState.Status.IN_PROGRESS
                messages.warning(
                    request,
                    'Check every item on the inspection protocol before marking this task complete. Your progress was saved.',
                )
        elif action == 'draft':
            if status_val != TaskState.Status.COMPLETED:
                status_val = TaskState.Status.IN_PROGRESS
        photo_n = TaskEvidencePhoto.objects.filter(derived_task_id=derived_id).count()
        TaskState.objects.update_or_create(
            derived_task_id=derived_id,
            defaults={
                'status': status_val,
                'notes': form.cleaned_data.get('notes') or '',
                'photo_count': photo_n,
                'checklist_done': done,
            },
        )
        if action == 'complete' and status_val == TaskState.Status.COMPLETED:
            messages.success(request, 'Task saved.')
            return redirect('workforce:worker_history')
        if not (action == 'complete' and status_val != TaskState.Status.COMPLETED):
            messages.success(request, 'Task saved.')
        if action == 'draft':
            return redirect('workforce:worker_my_tasks')
        return redirect('workforce:worker_task_detail', event_id=event_id, date_key=date_key)

    evidence_photos = list(TaskEvidencePhoto.objects.filter(derived_task_id=derived_id))
    pc = len(evidence_photos)

    display_status = effective_worker_display_status(
        st.status if st else TaskState.Status.SCHEDULED,
        checklist_keys,
        st.checklist_done if st else {},
    )

    return render(
        request,
        'workforce/worker/task_detail.html',
        {
            'event': ev,
            'task': task,
            'derived_id': derived_id,
            'form': form,
            'date_key': date_key,
            'persist': st,
            'display_status': display_status,
            'evidence_photos': evidence_photos,
            'photo_count': pc,
        },
    )
