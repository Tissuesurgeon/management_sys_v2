"""Parse maintenance import files — logic aligned with Management_sys parseMaintenanceImport.js."""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from workforce.models import CalendarEvent, Worker


def normalize_header(k: str) -> str:
    return re.sub(r'[\s#._-]+', '', str(k or '').lower())


def pick_field(row: Dict[str, Any], aliases: List[str]) -> Any:
    entries = [(k, v) for k, v in row.items() if v not in ('', None)]
    for a in aliases:
        na = normalize_header(a)
        for k, v in entries:
            if normalize_header(str(k)) == na:
                return v
    for a in aliases:
        na = normalize_header(a)
        if len(na) < 3:
            continue
        for k, v in entries:
            nk = normalize_header(str(k))
            if na in nk or nk in na:
                return v
    return None


def resolve_worker(raw: Any) -> Optional[Worker]:
    if raw is None or str(raw).strip() == '':
        return Worker.objects.order_by('pk').first()
    v = str(raw).strip()
    try:
        pk = int(v)
        w = Worker.objects.filter(pk=pk).first()
        if w:
            return w
    except (TypeError, ValueError):
        pass
    v_lower = v.lower()
    w = Worker.objects.filter(employee_id__iexact=v_lower).first()
    if w:
        return w
    w = Worker.objects.filter(name__iexact=v).first()
    if w:
        return w
    for x in Worker.objects.all():
        nl = x.name.lower()
        if v_lower in nl or (nl and nl.split()[0] in v_lower):
            return x
    return Worker.objects.order_by('pk').first()


def parse_recurrence(raw: Any) -> Dict[str, str]:
    if raw is None or str(raw).strip() == '':
        return {'type': 'none'}
    s = str(raw).strip().lower()
    if s in ('daily', 'day'):
        return {'type': 'daily'}
    if s in ('weekly', 'week'):
        return {'type': 'weekly'}
    if s in ('monthly', 'month'):
        return {'type': 'monthly'}
    return {'type': 'none'}


def parse_start(val: Any) -> datetime:
    if val is None or val == '':
        return datetime.now()
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, datetime.min.time())
    s = str(val).strip()
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s[:19] if len(s) > 10 else s, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace('Z', '+00:00'))
    except ValueError:
        pass
    m = re.match(r'^(\d{4}-\d{2}-\d{2})[ T](\d{1,2}):(\d{2})', s)
    if m:
        return datetime.strptime(
            f'{m.group(1)} {m.group(2).zfill(2)}:{m.group(3)}:00',
            '%Y-%m-%d %H:%M:%S',
        )
    m2 = re.match(r'^(\d{4}-\d{2}-\d{2})$', s)
    if m2:
        return datetime.strptime(m2.group(1) + ' 09:00:00', '%Y-%m-%d %H:%M:%S')
    return datetime.now()


def row_to_event_dict(row: Dict[str, Any], _index: int) -> Optional[Dict[str, Any]]:
    title = pick_field(
        row,
        ['title', 'task', 'activity', 'name', 'description', 'maintenance', 'workorder'],
    )
    if title is None or str(title).strip() == '':
        return None

    location = pick_field(row, ['location', 'site', 'zone', 'where', 'area', 'building']) or 'TBD'

    start_raw = pick_field(
        row,
        ['start', 'datetime', 'scheduled', 'when', 'starttime', 'startdate', 'timestamp'],
    )
    if start_raw in (None, ''):
        date_part = pick_field(row, ['date', 'day'])
        time_part = pick_field(row, ['time', 'clock', 'hour'])
        if date_part:
            start_raw = (
                f'{date_part} {time_part}'
                if time_part not in (None, '') and str(time_part).strip()
                else str(date_part)
            )
    if start_raw in (None, ''):
        start_raw = datetime.now()

    start = parse_start(start_raw)

    worker_raw = pick_field(
        row,
        ['worker', 'assignee', 'assigned', 'workerid', 'employee', 'technician'],
    )
    worker = resolve_worker(worker_raw)
    if not worker:
        return None

    recurrence = parse_recurrence(
        pick_field(row, ['recurrence', 'repeat', 'freq', 'frequency', 'interval']),
    )

    checklist_raw = pick_field(row, ['checklist', 'steps', 'items', 'steplist'])
    if checklist_raw not in (None, '') and str(checklist_raw).strip():
        checklist = [
            x.strip()
            for x in re.split(r'[|;]', str(checklist_raw))
            if x.strip()
        ]
    else:
        checklist = ['Verify completion']

    duration_raw = pick_field(row, ['duration', 'minutes', 'mins', 'length'])
    duration_minutes = 60
    if duration_raw not in (None, ''):
        try:
            n = int(float(duration_raw))
            if n > 0:
                duration_minutes = min(n, 24 * 60)
        except (TypeError, ValueError):
            pass

    color_raw = pick_field(row, ['color', 'tag'])
    color = 'blue'
    if color_raw is not None:
        c = str(color_raw).lower()
        if c in ('green', 'orange', 'red', 'blue'):
            color = c

    return {
        'title': str(title).strip(),
        'location': str(location).strip(),
        'start': start,
        'duration_minutes': duration_minutes,
        'worker': worker,
        'recurrence': recurrence,
        'checklist': checklist,
        'color': color,
    }


def rows_to_events(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    events: List[Dict[str, Any]] = []
    errors: List[str] = []
    if not rows:
        return [], ['No data rows found. Include a header row and at least one activity.']

    for i, row in enumerate(rows):
        try:
            ev = row_to_event_dict(row, i)
            if ev:
                events.append(ev)
        except Exception as e:
            errors.append(f'Row {i + 2}: {e!s}')

    if not events:
        errors.append(
            'No valid activities found. Required column: title (or task / activity). '
            'Optional: location, start, worker, recurrence, checklist.',
        )

    return events, errors


def parse_csv_text(text: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    rows = list(reader)
    return rows_to_events(rows)


def parse_xlsx_bytes(buf: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
    name = wb.sheetnames[0]
    sheet = wb[name]
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or '') for c in row]
            continue
        if not any(x is not None and str(x).strip() != '' for x in row):
            continue
        d = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        rows.append(d)
    wb.close()
    return rows_to_events(rows)


def parse_uploaded_file(name: str, content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    lower = name.lower()
    ext = lower[lower.rfind('.') :] if '.' in lower else ''

    if ext in ('.csv', '.txt'):
        return parse_csv_text(content.decode('utf-8', errors='replace'))

    if ext in ('.xlsx', '.xls'):
        return parse_xlsx_bytes(content)

    return [], [
        f'Unsupported format ({ext or "unknown"}). Use .csv, .xlsx, or .xls.',
    ]


def create_events_from_import(rows: List[Dict[str, Any]]) -> List[CalendarEvent]:
    created: List[CalendarEvent] = []
    for d in rows:
        rec = d['recurrence']
        rtype = rec.get('type', 'none')
        ev = CalendarEvent.objects.create(
            title=d['title'],
            location=d['location'],
            start=d['start'],
            duration_minutes=d['duration_minutes'],
            assigned_worker=d['worker'],
            recurrence_type=rtype,
            recurrence_end=None,
            checklist=d['checklist'],
            color=d['color'],
        )
        created.append(ev)
    return created
